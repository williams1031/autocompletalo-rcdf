import io
import re
import datetime
import base64
import json
import os
from typing import Dict

import streamlit as st
from openpyxl import load_workbook
from unidecode import unidecode

# =========================
# Configuración de página
# =========================
st.set_page_config(page_title="Autocomplétalo", page_icon="📄", layout="centered")
st.markdown("<div id='app-top'></div>", unsafe_allow_html=True)

# --------- Utilidades ---------
def get_base64_image(image_path: str) -> str:
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

logo = get_base64_image("logo_acueducto-01.jpg")

# Cargar estilos (si existen)
try:
    with open("styles/style.css") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
except FileNotFoundError:
    st.warning("⚠️ No se encontró el archivo de estilos (styles/style.css).")

# Encabezado
st.markdown(
    """
<div class='header'>
  <div>
    <h1>Autocomplétalo — Ing. William Rodríguez</h1>
    <p>Sistema inteligente de actualización de avisos T2</p>
  </div>
</div>
""",
    unsafe_allow_html=True,
)

st.write(
    "Sube tu archivo **.xlsx**, completa los campos y presiona **➕ Agregar código**. "
    "Usa **🧹 Limpiar** para vaciar el formulario (no borra la lista)."
)

# =========================
# Funciones Excel
# =========================
def norm(s: str) -> str:
    if s is None:
        return ""
    return unidecode(str(s)).strip().lower()

def find_sheet_case_insensitive(wb, sheet_name: str):
    target = norm(sheet_name)
    for name in wb.sheetnames:
        if norm(name) == target:
            return wb[name]
    for name in wb.sheetnames:
        if target in norm(name):
            return wb[name]
    return wb[wb.sheetnames[0]]

def find_row_by_code(ws, code_value) -> int:
    code_str = str(code_value).strip()
    code_int = None
    try:
        code_int = int(re.sub(r"\D", "", code_str))
    except Exception:
        pass
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell is None:
                continue
            if str(cell).strip() == code_str:
                return i
            if code_int is not None:
                try:
                    if int(cell) == code_int:
                        return i
                except Exception:
                    pass
    return -1

def build_header_index(ws) -> Dict[str, int]:
    header_row_idx = 1
    for r in range(1, 11):
        row = [c.value for c in ws[r]]
        if any(v is not None and str(v).strip() != "" for v in row):
            header_row_idx = r
            break
    headers = {}
    for col_idx, cell in enumerate(ws[header_row_idx], start=1):
        key = norm(cell.value)
        if key:
            headers[key] = col_idx
    return headers

def best_header_match(key: str, headers: Dict[str, int]) -> int:
    k = norm(key)
    if k in headers:
        return headers[k]
    for h, idx in headers.items():
        if k in h:
            return idx
    return -1

# =========================
# Opciones persistentes
# =========================
OPTIONS_FILE = "custom_options.json"

def load_custom_options():
    if os.path.exists(OPTIONS_FILE):
        with open(OPTIONS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_custom_options(options):
    with open(OPTIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(options, f, ensure_ascii=False, indent=2)

custom_options = load_custom_options()

# =========================
# Estado de la app
# =========================
if "registros" not in st.session_state:
    st.session_state["registros"] = []
if "confirmar_borrado" not in st.session_state:
    st.session_state["confirmar_borrado"] = False
if "form_version" not in st.session_state:
    st.session_state["form_version"] = 0

v = st.session_state["form_version"]

# =========================
# Entradas superiores
# =========================
archivo = st.file_uploader("📤 Sube tu archivo Excel (.xlsx)", type=["xlsx"])
nombre_hoja = st.text_input("📑 Nombre de la hoja dentro del archivo", value="PROGRAMACION")

def editable_select(label, options, base_key):
    if base_key not in custom_options:
        custom_options[base_key] = options.copy()
    opciones = custom_options[base_key]
    col1, col2 = st.columns([4, 1])
    with col2:
        quiero_otro = st.toggle("✏️ Otro", key=f"{base_key}_toggle_{v}", label_visibility="collapsed")
    with col1:
        sel = st.selectbox(label, [""] + opciones, key=f"{base_key}_sel_{v}", disabled=quiero_otro)
    manual = ""
    if quiero_otro:
        manual = st.text_input(
            f"Ingresar otro valor para '{label}'",
            key=f"{base_key}_txt_{v}",
            placeholder="Escribe aquí…",
        )
        if manual.strip() and manual.strip() not in opciones:
            opciones.append(manual.strip())
            custom_options[base_key] = opciones
            save_custom_options(custom_options)
    return manual.strip() if manual.strip() else sel.strip()

# =========================
# Formulario (estable)
# =========================
st.subheader("🧾 Datos del código")

with st.form("form_captura", clear_on_submit=False):
    colA, colB = st.columns(2)
    with colA:
        codigo = st.text_input("Código:", key=f"codigo_{v}")
        ejecuta = editable_select("Ejecuta:", ["EAAB", "CONTRATISTA"], "ejecuta")
        clase_uso = editable_select(
            "Clase de uso:",
            ["Comercial","Especial","Fachada en latas","Industrial","Multiusuario","N/A","No indica en el acta","Oficial","Residencial"],
            "clase"
        )
        ef_terreno = editable_select("Efectividad en terreno:", ["EFECTIVA","INEFECTIVA","DEPURADO"], "ef_terreno")
        ef_rcdf = editable_select("Efectividad para RCDF:", ["EFECTIVA","INEFECTIVA","DEPURADO","PARCIAL"], "ef_rcdf")
        visitas = editable_select("Visitas ejecutadas:", ["SI EJECUTADAS","NO EJECUTADAS"], "visitas")

    with colB:
        no_acta = st.text_input("No. Acta inspección:", key=f"no_acta_{v}")
        acta_firmada = editable_select("Acta firmada:", ["SI","NO"], "acta")
        actividad_economica = editable_select("Actividad económica:", ["N/A"], "actividad_economica")
        anomalia = editable_select("Anomalía/Causa inefectividad:", ["N/A"], "anomalia")
        porque_parcial = editable_select("Por que es parcial - Observación:", ["N/A"], "parcial")
        comunicacion = editable_select("Comunicación con usuario:", ["SI","NO"], "com")
        estado = editable_select("Estado:", ["GESTIONADO","EN GESTION"], "estado")

    fecha_cal = st.date_input("Fecha de ejecución:", value=datetime.date.today(), key=f"fecha_cal_{v}")
    fecha_ejecucion = fecha_cal.strftime("%d/%m/%Y")

    colBtns = st.columns(3)
    with colBtns[0]:
        submit = st.form_submit_button("➕ Agregar código", use_container_width=True)
    with colBtns[1]:
        limpiar = st.form_submit_button("🧹 Limpiar campos", use_container_width=True)
    with colBtns[2]:
        borrar = st.form_submit_button("🗑️ Borrar lista de registros", use_container_width=True)

    if submit:
        if not codigo.strip():
            st.error("Debes ingresar un código válido.")
        else:
            st.session_state["registros"].append({
                "Código": codigo,
                "No. Acta inspección": no_acta,
                "Acta firmada": acta_firmada,
                "Ejecuta": ejecuta,
                "Fecha de ejecución": fecha_ejecucion,
                "Actividad económica": actividad_economica,
                "Clase de uso": clase_uso,
                "Efectividad en terreno": ef_terreno,
                "Efectividad para RCDF": ef_rcdf,
                "Anomalía/Causa inefectividad": anomalia,
                "Por que es parcial - Observación": porque_parcial,
                "Comunicación": comunicacion,
                "Visitas ejecutadas": visitas,
                "Estado": estado,
            })
            st.success(f"✅ Código {codigo} agregado.")
            st.rerun()

    if limpiar:
        st.session_state["form_version"] += 1
        st.rerun()

    if borrar:
        st.session_state["confirmar_borrado"] = True
        st.rerun()

# Confirmación de borrado (persistente fuera del form)
if st.session_state["confirmar_borrado"]:
    st.warning("⚠️ ¿Seguro que deseas borrar **todos** los registros?")
    cA, cB = st.columns(2)
    with cA:
        if st.button("✅ Sí, borrar todo", key="borra_ok"):
            st.session_state["registros"] = []
            st.session_state["confirmar_borrado"] = False
            st.success("Lista de registros borrada.")
            st.rerun()
    with cB:
        if st.button("❌ Cancelar", key="borra_no"):
            st.session_state["confirmar_borrado"] = False
            st.info("Operación cancelada.")
            st.rerun()

# Tabla
if st.session_state["registros"]:
    st.markdown("---")
    st.subheader(f"📋 Registros agregados (total: {len(st.session_state['registros'])})")
    st.dataframe(st.session_state["registros"], use_container_width=True)

# =========================
# Procesar y descargar Excel
# =========================
if st.button("🧾 Procesar y descargar Excel", key="procesar"):
    if not archivo or not st.session_state["registros"]:
        st.error("Primero sube un archivo y agrega al menos un registro antes de procesar.")
    else:
        try:
            data = archivo.read()
            wb = load_workbook(io.BytesIO(data))
            ws = find_sheet_case_insensitive(wb, nombre_hoja)
            headers = build_header_index(ws)
            errores, ok = [], 0

            for registro in st.session_state["registros"]:
                codigo_val = registro.get("Código")
                row_idx = find_row_by_code(ws, codigo_val)
                if row_idx == -1:
                    errores.append(f"No se encontró el código {codigo_val}.")
                    continue
                for k, v in registro.items():
                    if k == "Código" or not v:
                        continue
                    col_idx = best_header_match(k, headers)
                    if col_idx != -1:
                        ws.cell(row=row_idx, column=col_idx, value=str(v))
                ok += 1

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

            st.success(f"✅ Excel actualizado correctamente con {ok} registro(s).")
            if errores:
                with st.expander("Ver detalles de errores"):
                    for e in errores:
                        st.write("•", e)

            st.download_button(
                label="⬇️ Descargar Excel actualizado",
                data=out,
                file_name="resultado_actualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"❌ Error al procesar el archivo: {e}")

# =========================
# Elementos fijos
# =========================
st.markdown(
    f"""
<img id="fixed-logo" src="data:image/png;base64,{logo}" alt="Logo Acueducto">
<div id="app-footer">💻 Desarrollado con ❤️ por <strong>William Rodríguez</strong> © 2025</div>

<!-- Botón info izquierda -->
<div id="info-btn">?</div>
<div id="info-tooltip">
    <b>Autocomplétalo RCDF</b><br>
    Versión: <b>2.5</b><br>
    Autor: William Rodríguez<br>
    Contacto: goss.williamr@gmail.com
</div>

<!-- Botón scroll -->
<a class="scroll-to-top" href="#app-top" title="Subir">
  <svg xmlns="http://www.w3.org/2000/svg" width="26" height="26" fill="white" viewBox="0 0 24 24">
    <path d="M12 4l-8 8h5v8h6v-8h5z"/>
  </svg>
</a>
""",
    unsafe_allow_html=True,
)
