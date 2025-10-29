import io
import re
import datetime
import base64
import json
import os
from typing import Dict
import streamlit as st
from openpyxl import load_workbook
from unidecode import unidecode

# ---------------------------------------------------------
# CONFIGURACIÓN INICIAL
# ---------------------------------------------------------
st.set_page_config(page_title="Autocomplétalo", page_icon="📄", layout="centered")
st.markdown("<div id='app-top'></div>", unsafe_allow_html=True)

# ==== FUNCIÓN PARA CARGAR EL LOGO ====
def get_base64_image(image_path):
    with open(image_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

logo = get_base64_image("logo_acueducto-01.jpg")

# ==== CARGAR ESTILOS ====
try:
    with open("styles/style.css") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
except FileNotFoundError:
    st.warning("⚠️ No se encontró el archivo de estilos (styles/style.css).")

# ==== ENCABEZADO SIN LOGO ====
st.markdown(f"""
<div class='header'>
    <div>
        <h1>Autocomplétalo — Ing. William Rodríguez</h1>
        <p>Sistema inteligente de actualización de avisos T2</p>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown(
    "Sube tu archivo **.xlsx**, completa los campos y presiona **➕ Agregar código**. "
    "Usa **🧹 Limpiar campos** para vaciar el formulario sin eliminar los registros cargados."
)

# ---------------------------------------------------------
# FUNCIONES DE UTILIDAD PARA EXCEL
# ---------------------------------------------------------
def norm(s: str) -> str:
    if s is None:
        return ""
    return unidecode(str(s)).strip().lower()

def find_sheet_case_insensitive(wb, sheet_name: str):
    target = norm(sheet_name)
    for name in wb.sheetnames:
        if norm(name) == target:
            return wb[name]
    for name in wb.sheetnames:
        if target in norm(name):
            return wb[name]
    return wb[wb.sheetnames[0]]

def find_row_by_code(ws, code_value) -> int:
    code_str = str(code_value).strip()
    code_int = None
    try:
        code_int = int(re.sub(r"\D", "", code_str))
    except Exception:
        pass
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell is None:
                continue
            if str(cell).strip() == code_str:
                return i
            if code_int is not None:
                try:
                    if int(cell) == code_int:
                        return i
                except Exception:
                    pass
    return -1

def build_header_index(ws) -> Dict[str, int]:
    header_row_idx = 1
    for r in range(1, 11):
        row = [c.value for c in ws[r]]
        if any(v is not None and str(v).strip() != "" for v in row):
            header_row_idx = r
            break
    headers = {}
    for col_idx, cell in enumerate(ws[header_row_idx], start=1):
        key = norm(cell.value)
        if key:
            headers[key] = col_idx
    return headers

def best_header_match(key: str, headers: Dict[str, int]) -> int:
    k = norm(key)
    if k in headers:
        return headers[k]
    for h, idx in headers.items():
        if k in h:
            return idx
    return -1

# ---------------------------------------------------------
# PERSISTENCIA DE OPCIONES PERSONALIZADAS
# ---------------------------------------------------------
OPTIONS_FILE = "custom_options.json"

def load_custom_options():
    if os.path.exists(OPTIONS_FILE):
        with open(OPTIONS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_custom_options(options):
    with open(OPTIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(options, f, ensure_ascii=False, indent=2)

custom_options = load_custom_options()

# ---------------------------------------------------------
# INTERFAZ PRINCIPAL
# ---------------------------------------------------------
archivo = st.file_uploader("📤 Sube tu archivo Excel (.xlsx)", type=["xlsx"])
nombre_hoja = st.text_input("📑 Nombre de la hoja dentro del archivo", value="PROGRAMACION")

st.session_state.setdefault("registros", [])
st.session_state.setdefault("confirmar_borrado", False)
st.session_state.setdefault("form_version", 0)
v = st.session_state["form_version"]

# ---------------------------------------------------------
# COMPONENTE SELECT CON OPCIÓN “OTRO”
# ---------------------------------------------------------
def editable_select(label, options, base_key):
    if base_key not in custom_options:
        custom_options[base_key] = options.copy()

    opciones = custom_options[base_key]
    col1, col2 = st.columns([4, 1])
    with col2:
        mostrar_otro = st.toggle("✏️ Otro",
                                 key=f"{base_key}_toggle_{v}",
                                 label_visibility="collapsed")
    with col1:
        seleccionado = st.selectbox(
            label,
            [""] + opciones,
            key=f"{base_key}_sel_{v}",
            disabled=mostrar_otro,
        )

    manual = ""
    if mostrar_otro:
        manual = st.text_input(
            f"Ingresar otro valor para '{label}'",
            key=f"{base_key}_txt_{v}",
            placeholder="Escribe aquí...",
        )
        if manual.strip() and manual.strip() not in opciones:
            opciones.append(manual.strip())
            custom_options[base_key] = opciones
            save_custom_options(custom_options)

    return manual.strip() if manual.strip() else seleccionado.strip()

# ---------------------------------------------------------
# FORMULARIO PRINCIPAL
# ---------------------------------------------------------
st.subheader("🧾 Datos del código")

codigo = st.text_input("Código:", key=f"codigo_{v}")
no_acta = st.text_input("No. Acta inspección:", key=f"no_acta_{v}")

acta_firmada = editable_select("Acta firmada:", ["SI", "NO"], "acta")
ejecuta = editable_select("Ejecuta:", ["EAAB", "CONTRATISTA"], "ejecuta")

fecha_cal = st.date_input("Fecha de ejecución:", value=datetime.date.today(), key=f"fecha_cal_{v}")
fecha_ejecucion = fecha_cal.strftime("%d/%m/%Y")

clase_uso = editable_select(
    "Clase de uso:",
    ["Comercial", "Especial", "Fachada en latas", "Industrial", "Multiusuario",
     "N/A", "No indica en el acta", "Oficial", "Residencial"],
    "clase"
)

actividad_economica = editable_select("Actividad económica:", ["N/A"], "actividad_economica")
ef_terreno = editable_select("Efectividad en terreno:", ["EFECTIVA", "INEFECTIVA", "DEPURADO"], "ef_terreno")
ef_rcdf = editable_select("Efectividad para RCDF:", ["EFECTIVA", "INEFECTIVA", "DEPURADO", "PARCIAL"], "ef_rcdf")
anomalia = editable_select("Anomalía/Causa inefectividad:", ["N/A"], "anomalia")
porque_parcial = editable_select("Por que es parcial - Observación:", ["N/A"], "parcial")
comunicacion = editable_select("Comunicación con usuario:", ["SI", "NO"], "com")
visitas = editable_select("Visitas ejecutadas:", ["SI EJECUTADAS", "NO EJECUTADAS"], "visitas")
estado = editable_select("Estado:", ["GESTIONADO", "EN GESTION"], "estado")

# ---------------------------------------------------------
# BOTONES DE ACCIÓN
# ---------------------------------------------------------
c1, c2, c3 = st.columns([1, 1, 1])
with c1:
    if st.button("➕ Agregar código"):
        if not codigo.strip():
            st.error("Debes ingresar un código válido.")
        else:
            registro = {
                "Código": codigo,
                "No. Acta inspección": no_acta,
                "Acta firmada": acta_firmada,
                "Ejecuta": ejecuta,
                "Fecha de ejecución": fecha_ejecucion,
                "Actividad económica": actividad_economica,
                "Clase de uso": clase_uso,
                "Efectividad en terreno": ef_terreno,
                "Efectividad para RCDF": ef_rcdf,
                "Anomalía/Causa inefectividad": anomalia,
                "Por que es parcial - Observación": porque_parcial,
                "Comunicación": comunicacion,
                "Visitas ejecutadas": visitas,
                "Estado": estado,
            }
            st.session_state["registros"].append(registro)
            st.success(f"✅ Código {codigo} agregado correctamente.")

with c2:
    if st.button("🧹 Limpiar campos"):
        st.session_state["form_version"] += 1
        st.rerun()
with c3:
    if st.button("🗑️ Borrar lista de registros"):
        st.session_state["confirmar_borrado"] = True

# ---------------------------------------------------------
# PROCESAR Y DESCARGAR EXCEL
# ---------------------------------------------------------
if st.button("🧾 Procesar y descargar Excel"):
    if not archivo or not st.session_state["registros"]:
        st.error("Primero sube un archivo y agrega al menos un registro antes de procesar.")
    else:
        try:
            data = archivo.read()
            wb = load_workbook(io.BytesIO(data))
            ws = find_sheet_case_insensitive(wb, nombre_hoja)
            headers = build_header_index(ws)
            errores, ok = [], 0

            for registro in st.session_state["registros"]:
                codigo = registro.get("Código")
                row_idx = find_row_by_code(ws, codigo)
                if row_idx == -1:
                    errores.append(f"No se encontró el código {codigo}.")
                    continue
                for k, v in registro.items():
                    if k == "Código" or not v:
                        continue
                    col_idx = best_header_match(k, headers)
                    if col_idx != -1:
                        ws.cell(row=row_idx, column=col_idx, value=str(v))
                ok += 1

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

            st.success(f"✅ Excel actualizado correctamente con {ok} registro(s).")
            if errores:
                with st.expander("Ver detalles de errores"):
                    for e in errores:
                        st.write("•", e)

            st.download_button(
                label="⬇️ Descargar Excel actualizado",
                data=out,
                file_name="resultado_actualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # Mostrar resumen por fecha
            st.markdown("---")
            st.subheader("📅 Códigos registrados por fecha")
            registros_por_fecha = {}
            for reg in st.session_state["registros"]:
                fecha = reg.get("Fecha de ejecución", "Sin fecha")
                if fecha not in registros_por_fecha:
                    registros_por_fecha[fecha] = []
                registros_por_fecha[fecha].append(reg)

            for fecha, regs in sorted(registros_por_fecha.items(), reverse=True):
                st.markdown(f"### 📆 {fecha}")
                for r in regs:
                    st.markdown(f"- {r['Código']}")

        except Exception as e:
            st.error(f"❌ Error al procesar el archivo: {e}")

# ---------------------------------------------------------
# ELEMENTOS FIJOS: LOGO, FOOTER, BOTONES
# ---------------------------------------------------------
st.markdown(f"""
<img id="fixed-logo" src="data:image/png;base64,{logo}" alt="Logo Acueducto">
<div id="app-footer">💻 Desarrollado con ❤️ por <strong>William Rodríguez</strong> © 2025</div>

<!-- BOTÓN INFO IZQUIERDA -->
<div id="info-btn">?</div>
<div id="info-tooltip">
    <b>Autocomplétalo RCDF</b><br>
    Versión: <b>2.5</b><br>
    Autor: William Rodríguez<br>
    Contacto: goss.williamr@gmail.com
</div>

<!-- BOTÓN SCROLL -->
<a class="scroll-to-top" href="#app-top" title="Subir">
    <svg xmlns="http://www.w3.org/2000/svg" width="26" height="26" fill="white" viewBox="0 0 24 24">
        <path d="M12 4l-8 8h5v8h6v-8h5z"/>
    </svg>
</a>
""", unsafe_allow_html=True)
    